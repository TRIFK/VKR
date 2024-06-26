from django.shortcuts import redirect
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.models import User, Group
from .forms import RegistrationForm
from django.contrib.auth.forms import AuthenticationForm
from .models import Product, ProductLocation, Supply, SupplyProduct, Shipment, ShipmentProduct, Order, OrderProduct, ProductType
from django.http import JsonResponse
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as OpenpyxlImage
from io import BytesIO
import qrcode
from django.http import HttpResponse
from django.views.decorators.http import require_POST, require_GET
from django.core.cache import cache
from django.contrib import messages
from .forms import OrderForm
import json
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import get_object_or_404, render
from datetime import timedelta
from django.utils import timezone




def welcome(request):
    messages.info(request, 'Ваша сессия истекла. Пожалуйста, войдите в систему снова.')
    redirect('/welcome/')
    return render(request, 'welcome.html')

def auth(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('home')
            else:
                return render(request, 'auth.html', {'form': form, 'error': 'Неверные учетные данные'})
    else:
        form = AuthenticationForm()
    return render(request, 'auth.html', {'form': form})

def in_group(group_name):
    def check(user):
        return user.groups.filter(name=group_name).exists() or user.is_superuser
    return check

def logout_view(request):
    logout(request)
    return redirect('welcome')


def register(request):
    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        if form.is_valid():
            try:
                user = form.save(commit=False)
                user.save()
                group = form.cleaned_data.get('group')
                user.groups.add(group)
                user = authenticate(username=form.cleaned_data['username'], password=form.cleaned_data['password1'])
                if user is not None:
                    login(request, user)
                    return redirect('home')
                else:
                    return render(request, 'register.html', {'form': form, 'error': 'Ошибка аутентификации'})
            except Exception as e:
                return render(request, 'register.html', {'form': form, 'error': f'Ошибка при сохранении пользователя: {str(e)}'})
        else:
            groups = Group.objects.all()

            # Вывод ошибок формы для отладки
            print(form.errors)
            return render(request, 'register.html', {'form': form, 'error': 'Форма не валидна', 'groups': groups})
    else:
        form = RegistrationForm()

    groups = Group.objects.all()
    context = {'form': form, 'groups': groups}
    return render(request, 'register.html', context)

def home(request):
    orders = Order.objects.all()
    orders_products = OrderProduct.objects.all()

    # Попытка получить ближайшую дату отгрузки из кеша Redis
    nearest_shipment_date = cache.get('nearest_shipment_date')

    if not nearest_shipment_date:
        # Если дата не найдена в кеше, выполним запрос к базе данных
        try:
            nearest_shipment = Shipment.objects.filter(date_shipped__gte=date.today()).earliest('date_shipped')
            nearest_shipment_date = nearest_shipment.date_shipped
            # Сохраняем результат в кеш Redis на 15 минут
            cache.set('nearest_shipment_date', nearest_shipment_date, timeout=900)  # 900 секунд = 15 минут
        except Shipment.DoesNotExist:
            nearest_shipment_date = None

    context = {
        'orders': orders,
        'nearest_shipment_date': nearest_shipment_date,
    }
    return render(request, 'home.html', context)


def products(request):
    products = Product.objects.all()
    context = {
        'products': products
    }
    return render(request, 'products.html', context)

def placement(request):
    product_locations = ProductLocation.objects.all()
    context = {
        'product_locations': product_locations
    }
    return render(request, 'placement.html', context)

def supplies(request):
    supplies = Supply.objects.all()
    supplies_products = SupplyProduct.objects.all()
    context = {
        'supplies': supplies,
        'supplies_products': supplies_products
    }
    return render(request, 'supplies.html', context)

def shipment(request):
    shipments = Shipment.objects.all()
    shipment_products = ShipmentProduct.objects.all()
    context = {
        'shipments': shipments,
        'shipment_products': shipment_products
    }
    return render(request, 'shipment.html', context)



def generate_invoice(request, order_id=0):
    # Handle form submission if method is POST
    if request.method == 'POST':
        selected_orders_ids = request.POST.getlist('orders.id')

        if not selected_orders_ids:
            # Redirect or display an error message if no orders are selected
            return redirect('orders')  # Replace with your URL name for orders list

        orders = Order.objects.filter(id__in=selected_orders_ids)
    else:
        # Fetch all orders if no orders are selected (GET request)
        orders = Order.objects.all()

    # Create a new Excel workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = 'Заказ'

    # Define headers for the Excel sheet
    headers = ['Клиент', 'Продукция', 'Сумма заказа', 'Дата заказа', 'QR код']

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    data_font = Font(bold=False, color="000000")
    data_alignment = Alignment(horizontal="left", vertical="center")

    # Write headers into the first row of the sheet and apply styles
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Write order data into subsequent rows
    for row_num, order in enumerate(orders, 2):  # Start from row 2 for data rows
        products_list = ', '.join([str(op.product) for op in order.products.all()])
        data = [order.customer, products_list, order.summary, order.date_ordered]

        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = data_font
            cell.alignment = data_alignment

        # Generate a QR code for the order
        qr_data = f"Order ID: {order.id}, Customer: {order.customer}, Summary: {order.summary}"
        qr_img = qrcode.make(qr_data)

        # Save the QR code image to a BytesIO object
        qr_bytes = BytesIO()
        qr_img.save(qr_bytes)
        qr_bytes.seek(0)

        # Load the QR code image into an openpyxl image object
        qr_image = OpenpyxlImage(qr_bytes)
        qr_image.width = 100  # Set the width of the QR code image
        qr_image.height = 100  # Set the height of the QR code image

        # Add the QR code image to the worksheet
        ws.add_image(qr_image, f'E{row_num}')

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save the Excel file to a BytesIO object
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Set up the HTTP response
    response = HttpResponse(excel_file.read(), content_type='application/vnd.openpyxl.sheet')
    filename = 'selected_orders.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'

    return response

@require_GET
def orders(request):
    orders = Order.objects.all()
    orders_products = OrderProduct.objects.all()
    products = Product.objects.all()
    product_types = ProductType.objects.all()

    # Преобразуем данные о продуктах в JSON-совместимый формат
    products_data = list(products.values('id', 'name'))

    context = {
        'orders': orders,
        'orders_products': orders_products,
        'products': products_data,  # Передаем только данные о продуктах
        'product_types': product_types
    }
    return render(request, 'orders.html', context)


@csrf_exempt
def create_order(request):
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            data = json.loads(request.body)
            data['completed'] = False

            products = data.get('selected_products', [])

            form_data = {
                'customer': data.get('customer', ''),
                'summary': data.get('summary', 0),
                'date_ordered': data.get('date_ordered', ''),
            }

            # Сначала создаем заказ (Order)
            form = OrderForm(form_data)
            if form.is_valid():
                order_instance = form.save(commit=False)
                order_instance.completed = False  # Устанавливаем completed на False
                order_instance.save()

                ids_ord = []
                # Теперь добавляем продукты в заказ (OrderProduct)
                for product_data in products:
                    product_id = product_data['ID']
                    quantity = product_data['quantity']
                    order_product = OrderProduct.objects.create(product_id=product_id, quantity=quantity)
                    product_in_summary = Product.objects.get(id=product_id)
                    order_instance.summary += product_in_summary.price * quantity
                    order_instance.save()
                    ids_ord.append(order_product.id)

                order_instance.products.set(ids_ord)
                return JsonResponse({'success': True})
            else:
                errors = form.errors.as_json()
                return JsonResponse({'success': False, 'error': errors}, status=400)

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)



@csrf_exempt
@require_POST
def delete_orders(request):
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            data = json.loads(request.body)
            order_ids = data.get('order_id', [])
            if order_ids and isinstance(order_ids, list):
                # Удаление заказов из базы данных
                deleted_count, _ = Order.objects.filter(id__in=order_ids).delete()
                if deleted_count > 0:
                    return JsonResponse({'success': True, 'message': 'Заказы успешно удалены'}, status=200)
                else:
                    return JsonResponse({'success': False, 'message': 'Заказы не найдены'}, status=404)
            else:
                return JsonResponse({'success': False, 'message': 'Некорректные данные'}, status=400)
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Некорректный формат JSON'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    else:
        return JsonResponse({'success': False, 'message': 'Неправильный тип запроса'}, status=400)


@csrf_exempt
@require_POST
def edit_order(request):
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            data = json.loads(request.body)
            order_id = data.get('order_id')
            date_ordered = data.get('date_ordered')
            products = data.get('selected_products', [])

            if not order_id or not date_ordered or not products:
                return JsonResponse({'success': False, 'message': 'Некорректные данные'}, status=400)

            order = get_object_or_404(Order, id=order_id)
            order.date_ordered = date_ordered
            order.save()

            # Очистка текущих продуктов в заказе
            order.products.clear()

            for product_data in products:
                product_id = product_data['ID']
                quantity = product_data['quantity']
                product = get_object_or_404(Product, id=product_id)

                order_product = OrderProduct.objects.create(
                    order=order,
                    product=product,
                    quantity=quantity
                )
                order.products.add(order_product)

            order.save()

            return JsonResponse({'success': True, 'order': {
                'id': order.id,
                'date_ordered': order.date_ordered,
                'products': [{'name': op.product.name, 'quantity': op.quantity} for op in order.products.all()]
            }}, status=200)

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Некорректный формат JSON'}, status=400)
        except Order.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Заказ не найден'}, status=404)
        except Product.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Один из продуктов не найден'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    else:
        return JsonResponse({'success': False, 'message': 'Неправильный тип запроса'}, status=400)


@require_POST
def complete_shipment(request):
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        data = json.loads(request.body)
        shipment_id = data.get('shipment_id')

        try:
            shipment = Shipment.objects.get(id=shipment_id)
        except Shipment.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Shipment not found'}, status=404)

        # Получение списка id продуктов из отгрузки
        product_ids = shipment.products.values_list('product__id', flat=True)

        print(f"Отгрузка ID: {shipment_id}")
        print(f"Клиент: {shipment.customer}")
        print(f"Дата отгрузки: {shipment.date_shipped}")
        print(f"ID продуктов: {list(product_ids)}")

        # Фильтрация заказов по клиенту и дате
        orders = Order.objects.filter(
            customer=shipment.customer,
        ).distinct()
        for order in orders:
            orderProductsIds = list(order.products.values_list('product__id', flat=True))
            print(orderProductsIds == list(product_ids))
            if orderProductsIds == list(product_ids):
                order.completed = True
                order.save()
                break
            else:
                print("Не было найдено совпадений")
                return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)

        return JsonResponse({'success': True}, status=200)
    else:
        return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)



def shareOrder(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    customer = order.customer
    order_products = order.products.all()
    products_ids = []
    for product in order_products:
        shipment = ShipmentProduct.objects.create(product_id=product.product.id, quantity=product.quantity)
        products_ids.append(shipment.id)
        products = ShipmentProduct.objects.filter(id__in=products_ids)
        shipment_instance = Shipment.objects.create(customer=customer, date_shipped=timezone.now() + timedelta(days=3))
        shipment_instance.products.set(products)
        shipment_instance.save()

    return JsonResponse({"Status": 200})